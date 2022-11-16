from django.shortcuts import render, redirect
from django.contrib.auth import logout
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User
from django.contrib import messages
'''
from gettext import install
from unittest import result
import cv2
import face_recognition
'''
#import subprocess
from subprocess import check_output
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from .models import Directivo, Profesor, Alumno, Grupo, Horario, Asistencia, Justificante, Usuario, Asignatura
from .forms import DirectivoForm, ProfesorForm, AlumnoForm, GrupoForm, HorarioForm, AsistenciaForm, JustificanteForm, UsuarioForm, AsignaturaForm
# Create your views here.

def salir(request):
    logout(request)
    return redirect('login/')

def inicio(request):
    return render(request, 'paginas/inicio.html')

def inicioD(request):
    return render(request, 'usuarios/directivo/inicioD.html')

def inicioP(request):
    return render(request, 'usuarios/profesor/inicioP.html')

def inicioA(request):
    return render(request, 'usuarios/alumno/inicioA.html')

def directivos(request):
    directivos = Directivo.objects.all()
    return render(request, 'usuarios/directivo/indexD.html', {'directivos': directivos})

def crearDirectivo(request):
    formulario = DirectivoForm(request.POST or None)
    if formulario.is_valid():
        formulario.save()
        return redirect('directivos')
    return render(request, 'usuarios/directivo/crear.html', {'formulario': formulario})

def editarDirectivo(request, id):
    directivo = Directivo.objects.get(id=id)
    formulario = DirectivoForm(request.POST or None, instance=directivo)
    if formulario.is_valid():
        formulario.save()
        return redirect('directivos')
    return render(request, 'usuarios/directivo/editar.html', {'formulario': formulario, 'directivo': directivo})

def eliminarDirectivo(request, id):
    directivo = Directivo.objects.get(id=id)
    directivo.delete()
    return redirect('directivos')

def profesores(request):
    profesores = Profesor.objects.all()
    return render(request, 'usuarios/profesor/indexP.html', {'profesores': profesores})

def crearProfesor(request):
    formulario = ProfesorForm(request.POST or None)
    if formulario.is_valid():
        formulario.save()
        return redirect('profesores')
    return render(request, 'usuarios/profesor/crear.html', {'formulario': formulario})

def editarProfesor(request, id):
    profesor = Profesor.objects.get(id=id)
    formulario = ProfesorForm(request.POST or None, instance=profesor)
    if formulario.is_valid():
        formulario.save()
        return redirect('profesores')
    return render(request, 'usuarios/profesor/editar.html', {'formulario': formulario, 'profesor': profesor})

def eliminarProfesor(request, id):
    profesor = Profesor.objects.get(id=id)
    profesor.delete()
    return redirect('profesores')

def alumnos(request):
    alumnos = Alumno.objects.all()
    return render(request, 'usuarios/alumno/indexA.html', {'alumnos': alumnos})

def crearAlumno(request):
    formulario = AlumnoForm(request.POST or None, request.FILES or None)
    if formulario.is_valid():
        formulario.save()
        return redirect('alumnos')
    return render(request, 'usuarios/alumno/crear.html', {'formulario': formulario})

def editarAlumno(request, matricula):
    alumno = Alumno.objects.get(matricula=matricula)
    formulario = AlumnoForm(request.POST or None, request.FILES or None, instance=alumno)
    if formulario.is_valid():
        formulario.save()
        return redirect('alumnos')
    return render(request, 'usuarios/alumno/editar.html', {'formulario': formulario, 'alumno': alumno})

def eliminarAlumno(request, matricula):
    alumno = Alumno.objects.get(matricula=matricula)
    alumno.delete()
    return redirect('alumnos')

def grupos(request):
    grupos = Grupo.objects.all()
    return render(request, 'grupos/indexG.html', {'grupos': grupos})

def crearGrupo(request):
    formulario = GrupoForm(request.POST or None)
    if formulario.is_valid():
        formulario.save()
        return redirect('grupos')
    return render(request, 'grupos/crear.html', {'formulario': formulario})

def editarGrupo(request, id):
    grupo = Grupo.objects.get(id=id)
    formulario = GrupoForm(request.POST or None, instance=grupo)
    if formulario.is_valid():
        formulario.save()
        return redirect('grupos')
    return render(request, 'grupos/editar.html', {'formulario': formulario, 'grupo': grupo})

def eliminarGrupo(request, id):
    grupo = Grupo.objects.get(id=id)
    grupo.delete()
    return redirect('grupos')

def horarios(request):
    horarios = Horario.objects.all()
    return render(request, 'horarios/indexH.html', {'horarios': horarios})

def crearHorario(request): 
    formulario = HorarioForm(request.POST or None)
    if formulario.is_valid():
        formulario.save()
        return redirect('horarios')
    return render(request, 'horarios/crear.html', {'formulario': formulario})

def editarHorario(request, id):
    horario = Horario.objects.get(id=id)
    formulario = HorarioForm(request.POST or None, instance=horario)
    if formulario.is_valid():
        formulario.save()
        return redirect('horarios')
    return render(request, 'horarios/editar.html', {'formulario': formulario, 'horario': horario})

def eliminarHorario(request, id):
    horario = Horario.objects.get(id=id)
    horario.delete()
    return redirect('horarios')

def asistencias(request):
    asistencias = Asistencia.objects.all()
    return render(request, 'asistencias/indexAsis.html', {'asistencias': asistencias})

def crearAsistencia(request): 
    formulario = AsistenciaForm(request.POST or None)
    if formulario.is_valid():
        formulario.save()
        return redirect('asistencias')
    return render(request, 'asistencias/crear.html', {'formulario': formulario})

def editarAsistencia(request, id):
    asistencia = Asistencia.objects.get(id=id)
    formulario = AsistenciaForm(request.POST or None, instance=asistencia)
    if formulario.is_valid():
        formulario.save()
        return redirect('asistencias')
    return render(request, 'asistencias/editar.html', {'formulario': formulario, 'asistencia': asistencia})

def eliminarAsistencia(request, id):
    asistencia = Asistencia.objects.get(id=id)
    asistencia.delete()
    return redirect('asistencias')

def justificantes(request):
    justificantes = Justificante.objects.all()
    return render(request, 'justificantes/indexJ.html', {'justificantes': justificantes})

def crearJustificante(request):
    formulario = JustificanteForm(request.POST or None, request.FILES or None)
    if formulario.is_valid():
        formulario.save()
        return redirect('justificantes')
    return render(request, 'justificantes/crear.html', {'formulario': formulario})

def editarJustificante(request, id):
    justificante = Justificante.objects.get(id=id)
    formulario = JustificanteForm(request.POST or None, request.FILES or None, instance=justificante)
    if formulario.is_valid():
        formulario.save()
        return redirect('justificantes')
    return render(request, 'justificantes/editar.html', {'formulario': formulario, 'justificante': justificante})

def eliminarJustificante(request, id):
    justificante = Justificante.objects.get(id=id)
    justificante.delete()
    return redirect('justificantes')

def usuarios(request):
    usuarios = User.objects.all()
    return render(request, 'usuarios/usuario/indexU.html', {'usuarios': usuarios})

def crearUsuario(request):
    if request.method == 'POST':
        formulario = UsuarioForm(request.POST)
        if formulario.is_valid():
            username = formulario.cleaned_data['username']
            messages.success(request, f'Usuario {username} creado correctamente')
            formulario.save()
            return redirect('usuarios')
    else:
        formulario = UsuarioForm()
    return render(request, 'usuarios/usuario/crear.html', {'formulario': formulario})

def editarUsuario(request, id):
    usuario = User.objects.get(id=id)
    formulario = UsuarioForm(request.POST or None, instance=usuario)
    if formulario.is_valid():
        username = formulario.cleaned_data['username']
        messages.success(request, f'Usuario {username} modificado correctamente')
        formulario.save()
        return redirect('usuarios')
    return render(request, 'usuarios/usuario/editar.html', {'formulario': formulario, 'usuario': usuario})

def eliminarUsuario(request, id):
    usuario = User.objects.get(id=id)
    usuario.delete()
    return redirect('usuarios')

def asignaturas(request):
    asignaturas = Asignatura.objects.all()
    return render(request, 'asignaturas/indexAsig.html', {'asignaturas': asignaturas})

def crearAsignatura(request):
    formulario = AsignaturaForm(request.POST or None)
    if formulario.is_valid():
        formulario.save()
        return redirect('asignaturas')
    return render(request, 'asignaturas/crear.html', {'formulario': formulario})

def editarAsignatura(request, clave):
    asignatura = Asignatura.objects.get(clave=clave)
    formulario = AsignaturaForm(request.POST or None, instance=asignatura)
    if formulario.is_valid():
        formulario.save()
        return redirect('asignaturas')
    return render(request, 'asignaturas/editar.html', {'formulario': formulario, 'asignatura': asignatura})

def eliminarAsignatura(request, clave):
    asignatura = Asignatura.objects.get(clave=clave)
    asignatura.delete()
    return redirect('asignaturas')
'''
def abrirReconocimiento(request):
    #Imagen a comparar
    image = cv2.imread("Images\Alfonso.jpg")
    face_loc = face_recognition.face_locations(image)[0]
    #print("face_loc:", face_loc)
    face_image_encodings = face_recognition.face_encodings(image, known_face_locations=[face_loc])[0]

    cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)

    while True:
        ret, frame = cap.read()
        if ret == False: break
        frame = cv2.flip(frame, 1)

        face_locations = face_recognition.face_locations(frame)
        if face_locations != []:
            for face_location in face_locations:
                face_frame_encodings = face_recognition.face_encodings(frame, known_face_locations=[face_location])[0]
                result = face_recognition.compare_faces([face_frame_encodings], face_image_encodings)
                print("Result:", result)

                if result[0] == True:
                    text = "Alfonso"
                    color = (125, 220, 0)
                else:
                    text = "Desconocido"
                    color = (50, 50, 255)

                cv2.rectangle(frame, (face_location[3], face_location[2]), (face_location[1], face_location[2] + 30), 2)
                cv2.rectangle(frame, (face_location[3], face_location[0]), (face_location[1], face_location[2]), color, 2)
                cv2.putText(frame, text, (face_location[3], face_location[2] + 20), 2, 0.7, (255, 255, 255), 1)

        cv2.imshow("Frame", frame)
        k = cv2.waitKey(1)
        if k == 27 & 0xFF:
            break

    cap.release()
    cv2.destroyAllWindows()
'''
def reportes(request):
    return render(request, 'reportes/reporte.html')

def generarReporte(request):
    campo = request.GET.get('campo')
    query = Asistencia.objects.filter(status = campo)
    wb = Workbook()
    bandera = True
    controlador = 8
    cont = 1
    ws = wb.active
    
    ws.title = 'Hoja'+str(cont)
    for i in query:
        if bandera:
            ws = wb.active
            ws.title = 'Hoja'+str(cont)
            #bandera = False
        '''else:
            ws = wb.create_sheet('Hoja'+str(cont))
        '''
        #Crear imagen en la hoja
        ws.merge_cells('B1:B4')

        imag = Image('C:/Users/alfon/Desktop/Proyecto Estancia II/Sistema Asistencias/SystemAsistencias/AppUpemor/templates/reportes/img/logo.png')
        imag.width = 100
        imag.height = 100
        ws.add_image(imag, 'B1')


        #Crear titulo en la hoja
        ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C1'].font = Font(name='Calibri', bold=True, size=20)
        ws['C1'].border = Border(left = Side(border_style='thin', color='FF000000'), right = Side(border_style='thin'),
                                    top = Side(border_style='thin'), bottom = Side(border_style='thin'))
        ws['C1'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        ws['C1'] = 'Reporte de asistencia'

        #cambiar caracteristicas de las celdas
        ws.merge_cells('C1:G1')

        ws.row_dimensions[1].height = 30

        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20

        #Crear encabezados o cabeceras
        ws.row_dimensions[3].height = 20

        ws['B7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B7'].font = Font(name='Calibri', bold=True, size=10)
        ws['B7'].border = Border(left = Side(border_style='thin', color='FF000000'), right = Side(border_style='thin'),
                                    top = Side(border_style='thin'), bottom = Side(border_style='thin'))
        ws['B7'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['B7'] = 'Matricula'

        ws['C7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C7'].font = Font(name='Calibri', bold=True, size=10)
        ws['C7'].border = Border(left = Side(border_style='thin', color='FF000000'), right = Side(border_style='thin'),
                                    top = Side(border_style='thin'), bottom = Side(border_style='thin'))
        ws['C7'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['C7'] = 'Alumno'

        ws['D7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D7'].font = Font(name='Calibri', bold=True, size=10)
        ws['D7'].border = Border(left = Side(border_style='thin', color='FF000000'), right = Side(border_style='thin'),
                                    top = Side(border_style='thin'), bottom = Side(border_style='thin'))
        ws['D7'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['D7'] = 'Apellidos'

        ws['E7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['E7'].font = Font(name='Calibri', bold=True, size=10)
        ws['E7'].border = Border(left = Side(border_style='thin', color='FF000000'), right = Side(border_style='thin'),
                                    top = Side(border_style='thin'), bottom = Side(border_style='thin'))
        ws['E7'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['E7'] = 'Status'

        ws['F7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['F7'].font = Font(name='Calibri', bold=True, size=10)
        ws['F7'].border = Border(left = Side(border_style='thin', color='FF000000'), right = Side(border_style='thin'),
                                    top = Side(border_style='thin'), bottom = Side(border_style='thin'))
        ws['F7'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type='solid')
        ws['F7'] = 'Fecha'

        #Generar los datos en el reporte
        #ws['B'+str(controlador)].alignment = Alignment(horizontal='center', vertical='center')
        #ws.merge_cells('C1:G1')
        ws.cell(row=controlador, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=controlador, column=2).font = Font(name='Calibri', size=10)
        ws.cell(row=controlador, column=2).border = Border(left = Side(border_style='thin', color='FF000000'), right = Side(border_style='thin'),
                                    top = Side(border_style='thin'), bottom = Side(border_style='thin'))
        ws.cell(row=controlador, column=2).value = i.id_alumno.matricula


        ws.cell(row=controlador, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=controlador, column=3).font = Font(name='Calibri', size=10)
        ws.cell(row=controlador, column=3).border = Border(left = Side(border_style='thin', color='FF000000'), right = Side(border_style='thin'),
                                    top = Side(border_style='thin'), bottom = Side(border_style='thin'))
        ws.cell(row=controlador, column=3).value = i.id_alumno.nombre


        ws.cell(row=controlador, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=controlador, column=4).font = Font(name='Calibri', size=10)
        ws.cell(row=controlador, column=4).border = Border(left = Side(border_style='thin', color='FF000000'), right = Side(border_style='thin'),
                                    top = Side(border_style='thin'), bottom = Side(border_style='thin'))
        ws.cell(row=controlador, column=4).value = i.id_alumno.apellidoP + ' ' + i.id_alumno.apellidoM


        ws.cell(row=controlador, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=controlador, column=5).font = Font(name='Calibri', size=10)
        ws.cell(row=controlador, column=5).border = Border(left = Side(border_style='thin', color='FF000000'), right = Side(border_style='thin'),
                                    top = Side(border_style='thin'), bottom = Side(border_style='thin'))
        ws.cell(row=controlador, column=5).value = i.status


        ws.cell(row=controlador, column=6).alignment = Alignment(horizontal='center')
        ws.cell(row=controlador, column=6).font = Font(name='Calibri', size=10)
        ws.cell(row=controlador, column=6).border = Border(left = Side(border_style='thin', color='FF000000'), right = Side(border_style='thin'),
                                    top = Side(border_style='thin'), bottom = Side(border_style='thin'))
        ws.cell(row=controlador, column=6).value = i.fechaReg

        controlador += 1
        
        cont += 1

    #establecer el nombre de mi archivo 
    nombre_archivo = "Reporte.xlsx"
    #definir el tipo de respuesta que voy a tener
    response = HttpResponse(content_type="application/ms-excel")
    content = "attachment; filename = {0}".format(nombre_archivo)
    response['Content-Disposition'] = content
    wb.save(response)
    return response
    
'''
def backup(request):
    subprocess.call(['mysqldump', '--user=root', '--password=1234', '--host=localhost', 'asistencia', '>', 'C:/Users/Usuario/Desktop/backup.sql'])
    subprocess.Popen(['C:/Users/Usuario/Desktop/backup.sql'])

respaldo = subprocess.Popen(['C:/Users/Usuario/Desktop/backup.sql'], shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
stdout, stderr = respaldo.communicate()
print(stdout)
print(stderr)
'''
'''
def restaurar(request):
    subprocess.Popen("C:/wamp64/bin/mysql/mysql5.7.36/bin/mysqldump -h localhost -u [ MySQL user, p. root ] -c [ upemor ] > sqldump.sql", shell=True)
    return render(request, 'usuarios/directivo/inicioD.html')
'''

'''
def backup(request):
    subprocess.Popen("mysqldump -u root -p12345 victimas > /home/proyecto/backup.sql")
    subprocess.Popen("gzip -c /home/proyecto/backup.sql > /home/proyecto/backup.gz")
    dataf = open('/home/proyecto/backups/backup.gz', 'r')
    return HttpResponse(dataf.read(), mimetype='application/x-gzip')


solucion...
subprocess.Popen("mysqldump -u root -p12345 victimas > /home/proyecto/backup.sql", shell=True)

#2
En Django, puedes crear un respaldo de tu base de datos usando el comando manage.py dumpdata. Ejecuta el comando a continuación para respaldar todos los modelos de tu aplicación:

python manage.py dumpdata --exclude contenttypes --indent 4 > db.json

'''